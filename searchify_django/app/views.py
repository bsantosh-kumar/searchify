from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.generic import FormView
from .forms import *

import pytesseract    # ======= > Add
from PIL import Image

# 
from docx import Document
import nltk
import re
import openpyxl
from nltk.corpus import stopwords

# nltk.download('stopwords')
sw_nltk = stopwords.words('english')

def index(request) :
    return HttpResponse("Hello, world. You're at the index.")

def extractWordsFromImage(image) :
    txt = pytesseract.image_to_string(Image.open(image))
    return txt

def extractWordsFromText(path):
    l=[]
    m={}
    
    lines=path.readlines()
    for count, line in enumerate(lines) :
        line = str(line, 'utf-8')  # to remove b'' infront of byte strings
        l=line.split()
        for i in l :
            i=i.lower()
            if i not in sw_nltk :
                if m.get(i)==None :
                    m[i]=[]
                    m[i].append(count+1)
                else :
                    m[i].append(count+1)
    return m

def extractWordsFromDocx(path) :
    document=Document(path) 
    m={}
    count=0
    for p in document.paragraphs :
        count=count+1
        l=[l for l in p.text.split() if l.lower() not in sw_nltk]
        for i1 in l :
            if i1 not in sw_nltk :
                i1=i1.lower()
                if m.get(i1)==None :
                    m[i1]=[]
                    m[i1].append(count)
                else :
                    m[i1].append(count)
    return m

def extractWordsFromXlxs(path) :
    wb_obj=openpyxl.load_workbook(path)
    sheet_obj=wb_obj.active
    m={}
    for i in range(1,sheet_obj.max_row+1) :
        for j in range(1,sheet_obj.max_column+1) :
            cell_obj=sheet_obj.cell(row=i,column=j)
            if m.get(cell_obj.value)==None :
                m[cell_obj.value]=[]
                m[cell_obj.value].append([i,j])
            else :
                m[cell_obj.value].append([i,j])
    return m

def split_string(txt) :
    import re
    txt = txt.lower()
    txt = re.sub(r'[^\w\s]',' ',txt)
    txt = re.sub(r'\s+',' ',txt)
    return txt.split()
    
def remove_stopwords(txt) :
    txt = [w for w in txt if not w in sw_nltk]
    return list(set(txt))

class HomeView(FormView):
    form_class = UploadForm
    template_name = 'app/index.html'
    success_url = '/'
    
    # def form_valid(self, form):
    #     upload = self.request.FILES['file']
    #     print("upload = ",upload)
    #     file_type = upload.content_type
    #     print("file_type = ",file_type)
        
    #     words = []

    #     try :
    #         if(file_type.endswith(('jpeg','png'))):
    #             txt = extractWordsFromImage(upload)
    #             words = split_string(txt)
    #             words = remove_stopwords(words)

    #             print(words)
    #             print('image')

    #         elif(file_type.endswith('pdf')):
    #             print('pdf')

    #         elif(file_type.endswith('sheet')):
    #             print('xlxs')
    #             words_locs = extractWordsFromXlxs(upload)
    #             print(words_locs)

    #         elif(file_type.endswith('text/plain')):
    #             words_lines = extractWordsFromText(upload)
    #             print(words_lines)
    #             print('txt')

    #         elif(file_type.endswith('document')):
    #             words_paras = extractWordsFromDocx(upload)
    #             print(words_paras)
    #             print('docx')

    #         else :
    #             print("not known type")

    #     except Exception as e:
    #         print("file parsing error")
    #         print(e)


    #     return super().form_valid(form)

def processFile(request) :
    if request.method == 'POST' :
        upload = request.FILES['file']
        print("upload = ",upload)
        file_type = upload.content_type
        print("file_type = ",file_type)
        
        words = []

        try :
            if(file_type.endswith(('jpeg','png'))):
                txt = extractWordsFromImage(upload)
                words = split_string(txt)
                words = remove_stopwords(words)

                print(words)
                print('image')

            elif(file_type.endswith('pdf')):
                print('pdf')

            elif(file_type.endswith('sheet')):
                print('xlxs')
                words_locs = extractWordsFromXlxs(upload)
                print(words_locs)

            elif(file_type.endswith('text/plain')):
                words_lines = extractWordsFromText(upload)
                print(words_lines)
                print('txt')

            elif(file_type.endswith('document')):
                words_paras = extractWordsFromDocx(upload)
                print(words_paras)
                print('docx')

            else :
                print("not known type")

        except Exception as e:
            print("file parsing error")
            print(e)
    
    return HttpResponse('Parsed, go back for another file, <a href="/"> go back </a>')
